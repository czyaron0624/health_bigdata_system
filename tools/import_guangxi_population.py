# -*- coding: utf-8 -*-
"""
导入广西人口数据到 population_info。

当前表结构不包含 year 字段，因此本脚本按“最新快照”导入：
1. 各市 2023 年年末常住人口 -> region 总量记录
2. 广西 2023 年年龄构成 -> 广西年龄段记录

人口单位假定为“万人”，导入时转换为“人”。
"""

from pathlib import Path

from openpyxl import load_workbook
import mysql.connector


BASE_DIR = Path(__file__).resolve().parents[1]
CITY_FILE = BASE_DIR / "inputs" / "广西省各市年末常住人口（2014-2023年）.xlsx"
AGE_FILE = BASE_DIR / "inputs" / "广西常住人口年龄构成（2014-2023年）.xlsx"

DB_CONFIG = {
    "host": "localhost",
    "user": "root",
    "password": "rootpassword",
    "database": "health_db",
}

LATEST_YEAR_COLUMN = "2023年"
AGE_GROUP_MAPPING = {
    "0-14岁占总人口的比重": "0-14",
    "15-64岁占总人口的比重": "15-64",
    "65岁及以上占总人口的比重": "65+",
}


def to_people(value_in_ten_thousand):
    return int(round(float(value_in_ten_thousand) * 10000))


def load_city_population():
    wb = load_workbook(CITY_FILE, data_only=True)
    ws = wb.active
    headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    year_index = headers.index(LATEST_YEAR_COLUMN)

    rows = []
    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        region = row[0]
        value = row[year_index]
        if not region or value is None:
            continue
        people = to_people(value)
        rows.append({"region": str(region).strip(), "population_count": people})
        total += people

    rows.append({"region": "广西", "population_count": total})
    return rows, total


def load_age_distribution(total_population):
    wb = load_workbook(AGE_FILE, data_only=True)
    ws = wb.active
    headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    year_index = headers.index(LATEST_YEAR_COLUMN)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        indicator = row[0]
        percentage = row[year_index]
        age_group = AGE_GROUP_MAPPING.get(indicator)
        if not age_group or percentage is None:
            continue
        rows.append(
            {
                "region": "广西",
                "age_group": age_group,
                "population_count": int(round(total_population * float(percentage) / 100)),
            }
        )
    return rows


def import_population():
    city_rows, total_population = load_city_population()
    age_rows = load_age_distribution(total_population)
    regions_to_replace = [row["region"] for row in city_rows]

    conn = mysql.connector.connect(**DB_CONFIG)
    try:
        cursor = conn.cursor()

        delete_sql = (
            "DELETE FROM population_info "
            f"WHERE region IN ({', '.join(['%s'] * len(regions_to_replace))})"
        )
        cursor.execute(delete_sql, regions_to_replace)

        insert_sql = """
            INSERT INTO population_info (region, age_group, gender, population_count)
            VALUES (%s, %s, %s, %s)
        """

        for row in city_rows:
            cursor.execute(insert_sql, (row["region"], None, None, row["population_count"]))

        for row in age_rows:
            cursor.execute(insert_sql, (row["region"], row["age_group"], None, row["population_count"]))

        conn.commit()
        cursor.close()
    finally:
        conn.close()

    print(f"Imported {len(city_rows)} regional population rows and {len(age_rows)} age rows")
    print(f"Guangxi total population (2023 snapshot): {total_population}")


if __name__ == "__main__":
    import_population()
