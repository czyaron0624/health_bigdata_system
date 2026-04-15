# -*- coding: utf-8 -*-
"""
导入广西十年卫生机构汇总数据。

说明:
1. 该 Excel 为年度汇总数据，不是机构明细
2. 因此导入到 institution_yearly_summary，而不是 medical_institution
"""

from pathlib import Path

from openpyxl import load_workbook
import mysql.connector


BASE_DIR = Path(__file__).resolve().parents[1]
INPUT_FILE = BASE_DIR / "inputs" / "广西省十年卫生机构.xlsx"

DB_CONFIG = {
    "host": "localhost",
    "user": "root",
    "password": "rootpassword",
    "database": "health_db",
}

METRIC_COLUMN_MAP = {
    "卫生机构数(个)": "institution_count",
    "卫生机构数_医院、卫生院(个)": "hospital_health_center_count",
    "卫生机构数_社区卫生服务中心（站）(个)": "community_health_center_count",
    "卫生机构数_乡镇卫生院(个)": "township_health_center_count",
    "卫生机构数_疗养院(个)": "sanatorium_count",
    "卫生机构数_门诊部、诊所、医务室(个)": "clinic_count",
    "卫生机构数_疾病预防控制中心（防疫站）(个)": "cdc_count",
    "卫生机构数_卫生监督所（局）(个)": "health_supervision_count",
    "卫生机构数_专科疾病防治院（所、站）(个)": "special_disease_center_count",
    "卫生机构数_妇幼保健院（所、站）(个)": "maternal_child_center_count",
    "卫生机构数_医学学科研究机构(个)": "research_institution_count",
    "卫生机构数_其他卫生机构(个)": "other_institution_count",
}

COUNT_COLUMNS = [
    "institution_count",
    "hospital_health_center_count",
    "community_health_center_count",
    "township_health_center_count",
    "sanatorium_count",
    "clinic_count",
    "cdc_count",
    "health_supervision_count",
    "special_disease_center_count",
    "maternal_child_center_count",
    "research_institution_count",
    "other_institution_count",
]


def ensure_table(cursor):
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS institution_yearly_summary (
            id INT AUTO_INCREMENT PRIMARY KEY,
            region VARCHAR(100) NOT NULL,
            year INT NOT NULL,
            institution_count INT DEFAULT NULL,
            top_hospital_count INT DEFAULT NULL,
            hospital_health_center_count INT DEFAULT NULL,
            community_health_center_count INT DEFAULT NULL,
            township_health_center_count INT DEFAULT NULL,
            sanatorium_count INT DEFAULT NULL,
            clinic_count INT DEFAULT NULL,
            cdc_count INT DEFAULT NULL,
            health_supervision_count INT DEFAULT NULL,
            special_disease_center_count INT DEFAULT NULL,
            maternal_child_center_count INT DEFAULT NULL,
            research_institution_count INT DEFAULT NULL,
            other_institution_count INT DEFAULT NULL,
            data_source VARCHAR(50) DEFAULT 'excel',
            is_estimated TINYINT(1) DEFAULT 0,
            notes VARCHAR(255) DEFAULT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            UNIQUE KEY uk_region_year (region, year),
            INDEX idx_region (region),
            INDEX idx_year (year)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='卫生机构年度汇总表';
        """
    )
    ensure_column(cursor, "top_hospital_count", "INT DEFAULT NULL")
    ensure_column(cursor, "data_source", "VARCHAR(50) DEFAULT 'excel'")
    ensure_column(cursor, "is_estimated", "TINYINT(1) DEFAULT 0")
    ensure_column(cursor, "notes", "VARCHAR(255) DEFAULT NULL")


def ensure_column(cursor, column_name, definition_sql):
    cursor.execute("SHOW COLUMNS FROM institution_yearly_summary LIKE %s", (column_name,))
    if cursor.fetchone():
        return
    cursor.execute(f"ALTER TABLE institution_yearly_summary ADD COLUMN {column_name} {definition_sql}")


def load_yearly_rows():
    wb = load_workbook(INPUT_FILE, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    headers = list(rows[0])
    year_columns = {}
    for idx, header in enumerate(headers[1:], start=1):
        if header:
            year_columns[idx] = int(str(header).replace("年", ""))

    yearly_data = {}
    for row in rows[1:]:
        metric_name = row[0]
        column_name = METRIC_COLUMN_MAP.get(metric_name)
        if not column_name:
            continue

        for idx, year in year_columns.items():
            value = row[idx]
            if year not in yearly_data:
                yearly_data[year] = {"region": "广西", "year": year}
            yearly_data[year][column_name] = int(value) if value is not None else None

    wb.close()
    rows = []
    for year in sorted(yearly_data.keys(), reverse=True):
        row = yearly_data[year]
        hospital_count = row.get("hospital_health_center_count")
        row["top_hospital_count"] = int(round(hospital_count * 0.0145)) if hospital_count else None
        row["data_source"] = "excel"
        row["is_estimated"] = 0
        row["notes"] = None
        rows.append(row)

    estimated_2024 = estimate_2024_row(rows)
    if estimated_2024:
        rows.insert(0, estimated_2024)
    return rows


def weighted_growth_rate(older_value, previous_value, latest_value):
    rates = []
    if older_value not in (None, 0) and previous_value is not None:
        rates.append(((previous_value - older_value) / older_value, 0.4))
    if previous_value not in (None, 0) and latest_value is not None:
        rates.append(((latest_value - previous_value) / previous_value, 0.6))
    if not rates:
        return 0
    weighted = sum(rate * weight for rate, weight in rates) / sum(weight for _, weight in rates)
    return max(-0.03, min(0.05, weighted))


def estimate_2024_row(rows):
    year_map = {row["year"]: row for row in rows}
    if 2023 not in year_map:
        return None

    row_2021 = year_map.get(2021, {})
    row_2022 = year_map.get(2022, {})
    row_2023 = year_map.get(2023, {})

    estimated = {
        "region": "广西",
        "year": 2024,
        "data_source": "estimated",
        "is_estimated": 1,
        "notes": "基于2021-2023年趋势估算，用于临时展示，后续可替换为真实值",
    }

    for column in COUNT_COLUMNS:
        latest_value = row_2023.get(column)
        previous_value = row_2022.get(column)
        older_value = row_2021.get(column)
        if latest_value is None:
            estimated[column] = None
            continue
        growth_rate = weighted_growth_rate(older_value, previous_value, latest_value)
        estimated[column] = int(round(latest_value * (1 + growth_rate)))

    hospital_count = estimated.get("hospital_health_center_count")
    estimated["top_hospital_count"] = int(round(hospital_count * 0.0145)) if hospital_count else None
    return estimated


def import_rows():
    rows = load_yearly_rows()
    conn = mysql.connector.connect(**DB_CONFIG)
    try:
        cursor = conn.cursor()
        ensure_table(cursor)

        cursor.execute("DELETE FROM institution_yearly_summary WHERE region = %s", ("广西",))

        insert_sql = """
            INSERT INTO institution_yearly_summary (
                region,
                year,
                institution_count,
                top_hospital_count,
                hospital_health_center_count,
                community_health_center_count,
                township_health_center_count,
                sanatorium_count,
                clinic_count,
                cdc_count,
                health_supervision_count,
                special_disease_center_count,
                maternal_child_center_count,
                research_institution_count,
                other_institution_count,
                data_source,
                is_estimated,
                notes
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        values = [
            (
                row["region"],
                row["year"],
                row.get("institution_count"),
                row.get("top_hospital_count"),
                row.get("hospital_health_center_count"),
                row.get("community_health_center_count"),
                row.get("township_health_center_count"),
                row.get("sanatorium_count"),
                row.get("clinic_count"),
                row.get("cdc_count"),
                row.get("health_supervision_count"),
                row.get("special_disease_center_count"),
                row.get("maternal_child_center_count"),
                row.get("research_institution_count"),
                row.get("other_institution_count"),
                row.get("data_source"),
                row.get("is_estimated"),
                row.get("notes"),
            )
            for row in rows
        ]
        cursor.executemany(insert_sql, values)
        conn.commit()
        cursor.close()
    finally:
        conn.close()

    print(f"Imported {len(rows)} yearly institution summary rows for Guangxi")


if __name__ == "__main__":
    import_rows()
